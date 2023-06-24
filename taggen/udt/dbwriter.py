# @AUTHOR: DIOCAI
# DEVELOP TIME: 23/4/28 22:58

from openpyxl import Workbook

from taggen.udt.datablock import DataBlock
from taggen.udt.udt import UDT
from taggen.udt.s7struct import S7Struct
from taggen.udt.extractor import DBExtractor

DB_HEADER = ['前缀', '位号', '类型', '位号描述', '所属DB', '生成偏移量', 'DB名称', '描述', '可否生成', '设定值']
UDT_HEADER = ['UDT', '后缀', '类型', '偏移量', '默认值', '临时1', '临时2', '临时3', '临时4', '临时5', '临时6', '描述', '报警', '只读', '别名', '是否创建']
STRUCT_HEADER = ['所属结构', '后缀', '类型', '偏移量', '默认值', '临时1', '临时2', '临时3', '临时4', '临时5', '临时6', '描述', '报警', '只读']

HEADER = {
    DataBlock.__name__: DB_HEADER,
    UDT.__name__: UDT_HEADER,
    S7Struct.__name__: STRUCT_HEADER,
}


class DBWriter:
    """
    将提取的DB写入到excel

    例子：

    e = DBExtractor('文件目录')

    writer = DBWriter()

    writer.write_defn_to_excel(e)

    writer.save('excel保存地址')
    """
    _wb: Workbook

    def __init__(self):
        self._wb = Workbook()
        if 'Sheet' in self._wb.sheetnames:
            self._wb.remove(self._wb['Sheet'])

    def write_defn_to_excel(self, _extractor, filename=None):
        if isinstance(_extractor, DBExtractor):
            self.write_list_to_excel(_extractor.db, 'data_block')
            if _extractor.udt:
                self.write_list_to_excel(_extractor.udt, 'udt')
            if _extractor.struct:
                self.write_list_to_excel(_extractor.struct, 'struct')
            if filename:
                self.save(filename)
        else:
            raise ValueError

    def write_list_to_excel(self, list_s7struct: list[DataBlock | UDT | S7Struct], sheet_name: str = None):
        wb = self._wb
        ws = wb.create_sheet(sheet_name) if sheet_name else wb.active
        # ws.append(DB_HEADER)
        ws.append(HEADER[list_s7struct[0].__class__.__name__])
        if isinstance(list_s7struct[0], DataBlock):
            for data_list in list_s7struct:
                for data in data_list.data:
                    ws.append(data.csv_format())
        else:
            for data_list in list_s7struct:
                for data in data_list.data:
                    ws.append(data.udt_format())

    def save(self, filename):
        wb = self._wb
        # ws = wb['Sheet']
        # wb.remove(ws)
        wb.save(filename)
