# @AUTHOR: DIOCAI
# DEVELOP TIME: 23/4/30 17:20
import openpyxl
import re


def read_db_number(filename):
    db_numbers = []
    if filename.endswith('.xlsx'):
        db_numbers = _read_db_number_from_excel(filename)

    if db_numbers:
        return db_numbers


def _read_db_number_from_excel(filename):
    """从excel中读取DB信息，未完成"""
    wb = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    ws = wb.active

    db_props = []
    props_header = [cell.value for cell in ws[1]]
    props_len = len(props_header)
    for row_cells in ws[2:ws.max_row]:
        props = [cell.value for cell in row_cells if cell.column <= props_len]
        db_props.append(props)
    return db_props


def test():
    db_numbers = read_db_number(r"D:\工作资料\11-PYTHON测试\DB地址定义.xlsx")
    print(db_numbers)


if __name__ == '__main__':
    DB_START_PAT = re.compile(r'(?P<type>DATA_BLOCK|TYPE)\s"(?P<title>.*)"')
    DB_END_PAT = re.compile(r'END_(DAT_BLOCK|TYPE)')
    STRUCT_PAT = re.compile('(?P<title>.*): Struct(?:   // (?P<comment>.*))?')
    TAG_PAT = re.compile(r'(?P<title>\w+) (?:\{ (?P<props>(.*) := \'(.*)\')+} )?: (.*);(?:   // (?P<comment>.*))?')

    res = DB_START_PAT.search('   TYPE "test"')
    res = TAG_PAT.search('B1 { ExternalWritable := \'False\'} : Bool;   // B1')
    res = TAG_PAT.search('U1 { S7_SetPoint := \'False\'} : "用户数据类型_1";   // U1')
    if res:
        print(res.groupdict())
    # test()
