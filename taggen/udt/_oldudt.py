# @AUTHOR: DIOCAI
# DEVELOP TIME: 23/3/16 9:54
import math
import os
import openpyxl

from taggen import tagtype
from taggen import UDTTag

UDT_FIELD_CHINESE = (
    'UDT',
    '后缀',
    '偏移量',
    '类型',
    '报警',
    '只读',
    '描述',
)

UDT_FIELD_ENGLISH = [
    'udt_type',
    'name',
    'offset',
    'type',
    'alarm_state',
    'read_only',
    'comment',
]

UDT_FIELD_DICT = {fd_chs: fd_eng for fd_eng, fd_chs in zip(UDT_FIELD_ENGLISH, UDT_FIELD_CHINESE)}


class OldUDT:
    author: str
    name: str
    version: str

    def __init__(self, udt_type=None, author=None, name=None, version=None, length=0, struct=None):
        if struct is None:
            struct = []
        self.udt_type = udt_type
        self.author = author
        self.name = name
        self.version = version
        self.length = length
        self.struct = struct

    def __getitem__(self, item):
        return self.__getattribute__(item)

    def __setitem__(self, key, value):
        if key in self.__dict__:
            self.__setattr__(key, value)
        else:
            print('No such field: "{}" in {}'.format(key, self.__class__))


class UDTList:
    pass


# 获取UDT的工厂方法
def get_udt_from(filename: str):
    if filename.endswith('.txt'):
        return get_udt_from_txt(filename)
    elif filename.endswith('udt_src'):
        return get_udt_from_path(filename)
    elif filename.endswith('.xlsx'):
        return get_udt_from_excel(filename)
    else:
        raise ValueError(f"{filename}不是有效的地址")


def get_udt_from_txt(filename: str) -> OldUDT | None:
    if filename.endswith('.txt'):
        print('正在处理: {}'.format(filename))
        f = open(filename, 'r', encoding='utf-8')
        lines = f.readlines()

        while True:
            temp_line = lines.pop(0)
            if 'DATA_BLOCK' in temp_line:
                print('目前无法处理DB文件')
                return
            if 'TYPE' in temp_line:
                break

        udt_obj = OldUDT()
        struct = []
        # udt 类型
        udt_obj.udt_type = temp_line[temp_line.find('"') + 1:temp_line.rfind('"')].strip()

        udt_fd = list(udt_obj.__dict__)
        # print(udt_fd)

        # udt 作者 描述 版本号
        for i in range(1, 4):
            temp_line = lines.pop(0)
            # print('udt_fd[{}]: , new_udt[{}]'.format(i, udt_fd[i])
            udt_obj[udt_fd[i]] = temp_line[temp_line.find(':') + 1:temp_line.rfind('\n')].strip().lower()

        # 读变量
        offset = 0
        prev_type = ''
        while len(lines) > 0:
            temp_line = lines.pop(0)
            if 'END_TYPE' in temp_line:
                break
            elif ':' not in temp_line:
                continue

            new_tag = _txt_read_tag(temp_line)

            if tagtype.is_bool(prev_type):
                if tagtype.is_bool(new_tag.type):
                    if offset * 10 % 10 > 7:
                        offset = math.floor(offset) + 1
                else:
                    offset = math.floor(offset)
                    if offset % 2 == 1:
                        offset += 1
                    else:
                        offset += 2
            if tagtype.is_bool(new_tag.type):
                new_tag.offset = '%.1f' % offset
            else:
                new_tag.offset = '%d' % offset

            offset += tagtype.type_length(new_tag.type)
            prev_type = new_tag.type
            struct.append(new_tag)
            udt_obj.struct = struct

        # udt 长度
        if tagtype.is_bool(prev_type):
            offset = math.floor(offset)
            if offset % 2 == 1:
                offset += 1
            else:
                offset += 2
        udt_obj.length = offset

        f.close()
        print('处理完毕')
        return udt_obj
    else:
        print('无法处理: {}'.format(filename))

    return


def get_udt_from_path(filepath: str) -> dict[str, OldUDT]:
    udt_dict = {}
    for root, dirs, files in os.walk(filepath):
        for f in files:
            # print(f)
            temp_udt = get_udt_from_txt(filepath + '\\' + f)
            if temp_udt is not None:
                udt_dict[temp_udt.udt_type] = temp_udt
    return udt_dict


# 还没写完
def get_udt_from_excel(filename: str) -> dict[OldUDT]:
    udt_dict = {}
    wb = openpyxl.load_workbook(filename=filename, data_only=True)
    if 'UDT汇总' in wb.sheetnames:
        ws = wb['UDT汇总']
    elif '类型描述及偏移量' in wb.sheetnames:
        ws = wb['类型描述及偏移量']
    else:
        ws = wb.active

    udt_index = {cell.value: cell.column - 1 for cell in ws['1']}
    print(udt_index)

    for row in ws[2:ws.max_row]:
        udt_type = row[udt_index['UDT']].value
        if udt_type not in udt_dict:
            # print(udt_type)
            udt_obj = OldUDT(name=udt_type)
            udt_dict[udt_type] = udt_obj

        # print(udt_dict[udt_type])

        tag_type = tagtype.type_name(row[udt_index['类型']].value)
        # print(tag_type)
        if tagtype.is_bool(tag_type):
            offset = round(row[udt_index['偏移量']].value, 1)
        else:
            offset = int(row[udt_index['偏移量']].value)

        new_tag = UDTTag(name=row[udt_index['后缀']].value,
                         offset=offset,
                         tag_type=tag_type,
                         alarm_state=row[udt_index['报警']].value,
                         read_only=row[udt_index['只读']].value,
                         comment=row[udt_index['描述']].value)

        udt_obj.struct.append(new_tag)

    return udt_dict


def _txt_read_tag(temp_line: str) -> UDTTag:
    end_name_index = temp_line.find(':')
    start_type_index = end_name_index + 1
    end_type_index = temp_line.find(';')
    # 变量名
    tag_name = temp_line[0:end_name_index].strip()
    # 变量类型
    tag_type = tagtype.type_name(temp_line[start_type_index:end_type_index].strip())
    # 注释
    if '//' in temp_line:
        tag_comment = temp_line[temp_line.find('//') + 2: temp_line.find('\n')].strip()
    else:
        tag_comment = None
    return UDTTag(name=tag_name, tag_type=tag_type, comment=tag_comment)


def read_struct(temp_line: str):
    print(temp_line)
    pass


if __name__ == '__main__':
    new_udt = get_udt_from_txt('d:/check_var.txt')
    for udt_tag in new_udt.struct:
        print(udt_tag)
