# @AUTHOR: DIOCAI
# DEVELOP TIME: 23/4/28 17:06
import os
import re
import logging

import openpyxl

from taggen.udt.datablock import DataBlock
from taggen.udt.s7data import S7Data
from taggen.udt.s7struct import S7Struct
from taggen.udt.udt import UDT


def read_db_defn(filename, sheet_name=None):
    db_defn = {}
    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        # 从excel获取定义列表
        for row_cells in ws[2: ws.max_row]:
            db_defn[row_cells[0].value] = DBDefn(*[row_cells[i].value for i in range(0, 5)])
        wb.close()
    return db_defn


def update_extractor(db_defn, extractor):
    for title in db_defn:
        if db := extractor.get_db_by_name(title):
            db.db_number = db_defn[title].db_number
            db.generable = db_defn[title].generable
            db.prefix = db_defn[title].new_prefix


class DBExtractor:
    __line = None
    __lines = None
    _struct_dict: dict[S7Struct] = {}
    _udt_dict: dict[UDT] = {}
    _bin: list[DataBlock] = []
    _db: list[DataBlock] = []

    def __init__(self, filename: str = None, udt_filepath=None):
        self.filename = filename
        self._udt_filepath = udt_filepath
        self._setup()

    def _setup(self):
        # 如果是文件夹 优先处理udt
        if os.path.isdir(self.filename):
            for _, _, filenames in os.walk(self.filename):
                filenames.sort(key=_sort_udt)
                for _f in filenames:
                    self._parse(self.filename + _f)
        else:
            self._parse(self.filename)
        print(f'处理完成，共 {len(self._db)} 个DB\n')
        logging.info(f'处理完成，共 {len(self._db)} 个DB')
        [logging.warning(f'无法添加的DB如下{db.title}: {db.error}') for db in self.bin]

    def _parse(self, filename=None):
        filename = filename if filename else self.filename
        if filename.endswith('.db'):
            self.parse_file(filename, DataBlock)
        elif filename.endswith('.udt'):
            self.parse_file(filename, UDT)
        else:
            logging.warning(f'无法处理：{filename}')

    def parse_file(self, filename, _Class=DataBlock):
        _s7objs = None
        rel_type = _Class.rel_type()

        print('正在处理：{}'.format(filename))
        f = open(filename, 'r', encoding='utf-8')
        self.__lines = f.readlines()

        # 遍历文件内容
        while True:
            try:
                self.__line = self.__lines.pop(0)
            except IndexError:
                f.close()
                self._release()
                break

            if res := re.search(f'{rel_type} "(?P<title>.*)"', self.__line):
                _s7objs = _Class(title=res.groupdict()['title'])
            elif _s7data := self.parse_data(parent=_s7objs):
                _s7objs.append(_s7data)
            elif _s7struct := self.parse_struct(parent=_s7objs):
                _s7objs.append(_s7struct)
            elif re.search(f'END_{rel_type}', self.__line):
                if _s7objs.generable:
                    print(f'{rel_type}名称：\'{_s7objs.title}\' 数据个数：{_s7objs.size}')
                    if isinstance(_s7objs, DataBlock):
                        self._db.append(_s7objs)
                    elif isinstance(_s7objs, UDT):
                        self._udt_dict[_s7objs.title] = _s7objs
                else:
                    self._bin.append(_s7objs)
                    print(f"无法添加，{rel_type}名称：{_s7objs.title}")

    def parse_data(self, parent=None, line=None):
        line = line if line else self.__line
        if res := re.search(r'(.*) : (.*);(.*)', line):
            # 变量名
            data_name = self._get_name(res.group(1).strip())

            # 变量注释
            data_remark = self._get_remark(res.group(3).strip())

            # 变量类型
            data_type = res.group(2).strip()

            _s7data = S7Data(parent=parent, title=data_name, data_type=data_type, comment=data_remark)

            try:
                if _res := re.search(r'"(.*)"', data_type):
                    _s7data.length = self.get_udt_length(_res[1])
            except KeyError as e:
                _s7data.generable = False
                if _s7data.data_block:
                    _s7data.data_block.generable = False
                    _s7data.data_block.error.append(f'{_s7data.title} 无法获得 {e} 的长度')
                _s7data.length = 0

            return _s7data

    def parse_struct(self, parent):
        if res := re.search(r'(.*): (Struct.*)', self.__line):
            # 结构名
            st_name = self._get_name(res.group(1))
            # 结构注释
            st_remark = self._get_remark(res.group(2).strip())
            _struct = S7Struct(parent=parent, title=st_name, comment=st_remark)
            self.struct_dict[_struct.data_type] = _struct
            parent.data_block.struct_dict[_struct.data_type] = _struct
        else:
            return

        while not re.search(r'end_struct', self.__line, re.I):
            try:
                self.__line = self.__lines.pop(0)
            except IndexError:
                break

            # 结构里的数据
            if st_data := self.parse_data(parent=_struct):
                _struct.append(st_data)
            # 嵌套结构
            if child_struct := self.parse_struct(parent=_struct):
                _struct.append(child_struct)
        return _struct

    def read_udt_from_excel(self, filename):
        pass

    def read_db_number(self, filename, sheet_name=None):
        """读取db地址 excel格式:db名称 | db地址 | """
        db_numbers = {}
        if filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            # 从excel获取定义列表
            for row_cells in ws[2: ws.max_row]:
                db_numbers[row_cells[0].value] = row_cells[1].value
            wb.close()
        for db_title in db_numbers:
            if db := self.get_db_by_name(db_title):
                db.db_number = db_numbers[db_title]
        # print(self._db_numbers)

    @staticmethod
    def _get_remark(string: str):
        if res := re.search(r'//(.*)', string):
            return res.group(1).strip()
        else:
            return ''

    @staticmethod
    def _get_name(string: str):
        if res := re.search(r'(.*){(.*)}', string):
            return res.group(1).strip()
        else:
            return string.strip()

    def _release(self):
        self.__line = None
        self.__lines = None

    def get_udt_length(self, _type):
        return self._udt_dict[_type].length

    @property
    def db_dict(self) -> dict[DataBlock]:
        return {db.title: db for db in self._db}

    @property
    def db(self):
        return self._db

    @property
    def udt_dict(self) -> dict[UDT]:
        return self._udt_dict

    @property
    def udt(self) -> list[UDT]:
        return [self._udt_dict[st] for st in self._udt_dict]

    @property
    def struct_dict(self) -> dict[S7Struct]:
        return self._struct_dict

    @property
    def struct(self) -> list[S7Struct]:
        return [self._struct_dict[st] for st in self._struct_dict]

    @property
    def all_struct(self) -> dict[S7Struct]:
        __all_struct = {}
        __all_struct.update(self._struct_dict)
        __all_struct.update(self._udt_dict)
        return __all_struct

    @property
    def bin(self):
        return self._bin

    def get_db_by_name(self, name):
        try:
            return self.db_dict[name]
        except KeyError:
            pass


class DBDefn:
    def __init__(self, title, db_number, comment, new_prefix, generable):
        self.title = title
        self.db_number = db_number
        self.comment = comment
        self.new_prefix = new_prefix if new_prefix else title
        if isinstance(generable, bool):
            self.generable = generable
        else:
            self.generable = True if generable == '是' else False


def _sort_udt(elem: str):
    """
    将.udt类型的文件排在列表前端
    :param elem: 文件名称
    :return: 排序优先级
    """
    if elem.endswith('.udt'):
        return -1
    else:
        return 1
