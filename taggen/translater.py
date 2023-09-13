# @AUTHOR: DIOCAI
# DEVELOP TIME: 23/9/8 10:08
import csv
import re

import openpyxl

from taggen.tag import HMITag, TagList
import taggen.intouch
import taggen.kepserver

HEADER_MAPPING = {
    'Intouch变量名': 'name',
    '组态王变量名': 'comment',
    '所属分组': 'group',
    '设备名': 'device',
    '寄存器名称': 'register',
    '数据类型': 'data_type',
    '书写属性': 'readonly',
    '报警信息': 'alarm_comment'
}

HEADER_MAPPING_REV = {HEADER_MAPPING[key]: key for key in HEADER_MAPPING}

# 需要的字段
FIELD_NEEDED = ['变量ID', '变量名', '设备名', '寄存器名称', '数据类型', '读写属性', '注释']

DATATYPE_MAPPING = {
    'bit': 'bool',
    'short': 'int',
    'ushort': 'int',
    'float': 'real'
}

DATATYPE_SUFFIX_MAPPING = {
    'bit': 'X',
    'short': 'INT',
    'ushort': 'INT',
    'float': 'REAL'
}


class KingViewTag:
    def __init__(self, name, comment, group, device, register, data_type, readonly, alarm_comment):
        self.name = name
        self.comment = comment
        self.group = group
        self.device = device
        self.register = register
        self.data_type = data_type
        self.readonly = readonly
        self.alarm_comment = alarm_comment

    def __str__(self):
        info = [f'{HEADER_MAPPING_REV[key]}: {getattr(self, key)}' for key in self.__dict__]
        return '\t'.join(info)


def read_data_from_excel(filename, sheet_name=None):
    new_data = []
    if filename.endswith('.xlsx'):
        # 打开excel
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        # 打开的sheet
        ws = wb[sheet_name] if sheet_name else wb.active
        # 从第二行开始读取数据
        for row_cells in ws[2:ws.max_row]:
            new_data.append([cell.value for cell in row_cells])
        wb.close()

    return new_data


def read_group_from_excel(filename, sheet_name=None):
    group_mapping = {}
    if filename.endswith('.xlsx'):
        # 打开excel
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        # 打开的sheet
        ws = wb[sheet_name] if sheet_name else wb.active
        # 从第二行开始读取数据
        for row_cells in ws[2:ws.max_row]:
            group_mapping[row_cells[0].value] = row_cells[1].value
        wb.close()

    return group_mapping


def ktags_sorted_by_access(datas):
    access_dict = {}
    for data in datas:
        ktag = KingViewTag(*data)
        device = ktag.device
        if device not in access_dict:
            access_dict[device] = []
        access_dict[device].append(ktag)

    return access_dict


def kingview_to_hmitag(ktag: KingViewTag, group_mapping=None):
    new_dict = {}
    new_dict['name'] = ktag.name
    new_dict['comment'] = ktag.comment
    new_dict['data_type'] = translate_datatype(ktag.data_type)
    new_dict['group'] = translate_group(ktag.group, group_mapping)
    # new_dict['device'] = ktag.device
    new_dict['read_only'] = 'Yes' if ktag.readonly == '只读' else 'No'
    new_dict['item_name'] = translate_register(ktag.register, ktag.data_type)
    new_dict['alarm_state'] = 1 if ktag.alarm_comment else 0
    hmitag = HMITag(**new_dict)
    return hmitag


def translate_datatype(datatype):
    return DATATYPE_MAPPING[datatype.lower()]


def translate_register(register: str, datatype):
    first_dot_pos = 0

    # 寄存器类型
    if register.startswith('DB'):
        first_dot_pos = register.find('.')
        item_name = register[:first_dot_pos] + ','
    else:
        item_name = register[0]

    # 类型后缀
    datatype = datatype.lower()
    datatype_suffix = DATATYPE_SUFFIX_MAPPING[datatype]
    item_name += datatype_suffix

    # 偏移量
    item_name += register[first_dot_pos + 1:]
    return item_name


def translate_group(group, group_mapping=None):
    return group_mapping[group] if group_mapping else group


def sorted_taglist(hmitags, taglist):
    for hmitag in hmitags:
        if hmitag.data_type == 'bool':
            taglist.disc_list.append(hmitag)
        elif hmitag.data_type == 'int':
            taglist.int_list.append(hmitag)
        elif hmitag.data_type == 'real':
            taglist.real_list.append(hmitag)
        else:
            raise ValueError(f'无法处理{hmitag.data_type}')


def read_kingview_db(filename, sheet_name='基本变量页'):
    datas = {}
    if filename.endswith('.xlsx'):
        # 打开excel
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        # 打开的sheet
        ws = wb[sheet_name] if sheet_name else wb.active

        row_cells = iter(ws[2:ws.max_row])
        newdtype = False
        while True:
            try:
                row = next(row_cells)
                cells = row
                row0_value = str(row[0].value)
                if res := re.search(r'\[IO(\w{2})\]', row0_value):  # 找到新的类型，只要IO类型
                    newdtype = True
                    dtype = res.group(0)[1:-1]
                    datas[dtype] = []
                    print(dtype)
                elif row0_value == '变量ID':  # 收集类型的字段
                    fd = {}
                    for cell in cells:
                        try:
                            fd[cell.column] = cell.value  # 收集所有字段
                        except AttributeError:
                            break
                elif newdtype:
                    try:
                        int(row0_value)  # 由于变量ID是数字，通过try来验证第一列是不是变量ID，不是则证明该类型已结束
                        d = {}
                        for cell in cells:
                            try:
                                # 只采集需要的字段
                                f = fd[cell.column]
                                if f in FIELD_NEEDED:
                                    d[f] = cell.value
                            except AttributeError:
                                # print(d)
                                continue
                        print(d)
                        datas[dtype].append(d)
                    except ValueError:
                        print(len(datas[dtype]))
                        newdtype = False

            except StopIteration:
                break

        wb.close()
    return datas


def write_kingview_csv(output_path, datas):
    with open(output_path, 'w', newline='', encoding='ANSI') as f:
        writer = csv.writer(f)
        writer.writerow(FIELD_NEEDED)

        for data_group in datas.values():
            for data in data_group:
                writer.writerow(data.values())


def translate_kingview_db(db_xlsx, output_csv):
    sortd_datas = read_kingview_db(db_xlsx)
    write_kingview_csv(output_csv, sortd_datas)


if __name__ == '__main__':
    xlsx_file = r'D:\工作资料\09-待分类\南靖中福变量改造对照表.xlsx'
    group_mapping = read_group_from_excel(xlsx_file, '分组')
    datas = read_data_from_excel(xlsx_file, '变量列表')
    ktag_access_dict = ktags_sorted_by_access(datas)

    # 根据设备分类
    hmitag_access_dict = {}
    for access in ktag_access_dict:
        # print(access)
        if access not in hmitag_access_dict:
            hmitag_access_dict[access] = []

        for ktag in ktag_access_dict[access]:
            hmitag = kingview_to_hmitag(ktag, group_mapping)
            hmitag_access_dict[access].append(hmitag)

    #
    taglist_dict = {}
    for access in hmitag_access_dict:
        # print(len(hmitag_access_dict[access]))
        if access not in taglist_dict:
            taglist_dict[access] = TagList()
        sorted_taglist(hmitag_access_dict[access], taglist_dict[access])

    # 分设备生成
    for access in taglist_dict:
        print(f'\ntaglist {access}存在: {len(taglist_dict[access])}个数据')
        taglist = taglist_dict[access]
        intouch_output_path = f'D:\\工作资料\\09-待分类\\intouch_{access}_db.csv'
        kep_output_path = f'D:\\工作资料\\09-待分类\\kepserver_{access}_db.csv'
        taggen.intouch.output_csv(taglist, intouch_output_path, mode='replace', access_name=access, item_use_tag_name=True)
        taggen.kepserver.output_csv(taglist, kep_output_path)

    # db_xlsx = r'D:\工作资料\09-待分类\ztw_db.xlsx'
    # output_csv = r'D:\工作资料\09-待分类\ztw_db.csv'
    # translate_kingview_db(db_xlsx, output_csv)
